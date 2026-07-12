from io import BytesIO
import csv
import calendar
from datetime import date
from decimal import Decimal
import pandas as pd
from django.db.models import Sum
from django.http import HttpResponse
from django.contrib.auth.models import User

# ReportLab imports
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas
from reportlab.graphics.shapes import Drawing, Rect, String as DString

# openpyxl imports for Excel
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.response import Response
from budgets.models import Budget
from transactions.models import Transaction


def transaction_frame(user, month=None, year=None):
    query = Transaction.objects.filter(user=user)
    if month:
        query = query.filter(date__month=month)
    if year:
        query = query.filter(date__year=year)
    return pd.DataFrame(list(query.values("date", "category", "transaction_type", "amount", "description")))


def currency(value):
    return f"Rs. {float(value or 0):,.2f}"


def financial_insights(user, month, year):
    current = Transaction.objects.filter(user=user, date__month=month, date__year=year)
    expenses = current.filter(transaction_type="expense")
    total_expenses = expenses.aggregate(total=Sum("amount"))["total"] or 0
    income = current.filter(transaction_type="income").aggregate(total=Sum("amount"))["total"] or 0
    insights = []
    categories = list(expenses.values("category").annotate(total=Sum("amount")).order_by("-total"))
    if categories and total_expenses:
        top = categories[0]
        insights.append(f"Your highest spending category this month was {top['category']} ({currency(top['total'])}), accounting for {float(top['total'] / total_expenses * 100):.0f}% of total expenses.")
    previous_month, previous_year = (month - 1, year) if month > 1 else (12, year - 1)
    previous = Transaction.objects.filter(user=user, date__month=previous_month, date__year=previous_year)
    previous_income = previous.filter(transaction_type="income").aggregate(total=Sum("amount"))["total"] or 0
    previous_expenses = previous.filter(transaction_type="expense").aggregate(total=Sum("amount"))["total"] or 0
    if income and previous_income:
        current_rate = float((income - total_expenses) / income * 100)
        previous_rate = float((previous_income - previous_expenses) / previous_income * 100)
        change = current_rate - previous_rate
        insights.append(f"Your savings rate {'improved' if change >= 0 else 'declined'} by {abs(change):.0f}% compared to the previous month.")
    entertainment = expenses.filter(category__iexact="Entertainment").aggregate(total=Sum("amount"))["total"] or 0
    previous_entertainment = previous.filter(transaction_type="expense", category__iexact="Entertainment").aggregate(total=Sum("amount"))["total"] or 0
    if entertainment and previous_entertainment:
        change = float((entertainment - previous_entertainment) / previous_entertainment * 100)
        insights.append(f"Entertainment spending {'increased' if change >= 0 else 'decreased'} by {abs(change):.0f}% compared to the previous month.")
    for budget in Budget.objects.filter(user=user, month=month, year=year):
        spent = expenses.filter(category__iexact=budget.category).aggregate(total=Sum("amount"))["total"] or 0
        if spent > budget.monthly_limit:
            insights.append(f"You exceeded your {budget.category} budget by {currency(spent - budget.monthly_limit)}.")
    if total_expenses:
        days = date.today().day if (month == date.today().month and year == date.today().year) else calendar.monthrange(year, month)[1]
        insights.append(f"Your average daily expense is {currency(total_expenses / days)}.")
    largest = current.order_by("-amount").first()
    if largest:
        insights.append(f"Your largest transaction was {largest.category} {currency(largest.amount)}.")
    return insights or ["Add transactions to receive personalised financial insights."]


class MonthlyReportView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        today = date.today()
        month = int(request.query_params.get("month", today.month))
        year = int(request.query_params.get("year", today.year))
        qs = Transaction.objects.filter(user=request.user, date__month=month, date__year=year)
        totals = qs.values("transaction_type").annotate(total=Sum("amount"))
        amounts = {row["transaction_type"]: row["total"] for row in totals}
        category = qs.filter(transaction_type="expense").values("category").annotate(total=Sum("amount")).order_by("-total")
        return Response({
            "month": month,
            "year": year,
            "income": amounts.get("income", 0),
            "expenses": amounts.get("expense", 0),
            "savings": amounts.get("income", 0) - amounts.get("expense", 0),
            "by_category": list(category),
            "insights": financial_insights(request.user, month, year)
        })


class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        # Header (pages > 1)
        if self._pageNumber > 1:
            self.setFont("Helvetica-Bold", 8)
            self.setFillColor(colors.HexColor("#64748b"))
            self.drawString(54, 755, "LEDGERLY — PERSONAL FINANCE STATEMENT")
            self.setStrokeColor(colors.HexColor("#e2e8f0"))
            self.setLineWidth(0.5)
            self.line(54, 747, letter[0] - 54, 747)

        # Footer (all pages)
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#64748b"))
        self.drawString(54, 30, "Generated by Ledgerly | Confidential")
        self.setStrokeColor(colors.HexColor("#e2e8f0"))
        self.setLineWidth(0.5)
        self.line(54, 42, letter[0] - 54, 42)

        page_str = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(letter[0] - 54, 30, page_str)
        self.restoreState()


def make_progress_bar(percentage):
    d = Drawing(80, 10)
    # background gray bar
    d.add(Rect(0, 2, 80, 6, fillColor=colors.HexColor("#e2e8f0"), strokeColor=None))
    # progress colored bar
    fill_color = colors.HexColor("#ef4444") if percentage >= 100 else colors.HexColor("#f59e0b") if percentage >= 80 else colors.HexColor("#10b981")
    width = min(80, max(0, float(percentage) / 100 * 80))
    if width > 0:
        d.add(Rect(0, 2, width, 6, fillColor=fill_color, strokeColor=None))
    return d


class ExportReportView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, export_format):
        today = date.today()
        month_param = request.query_params.get("month")
        year_param = request.query_params.get("year")
        
        month = int(month_param) if month_param else today.month
        year = int(year_param) if year_param else today.year
        month_name = calendar.month_name[month]

        # Get data
        qs = Transaction.objects.filter(user=request.user, date__month=month, date__year=year)
        total_income = qs.filter(transaction_type="income").aggregate(total=Sum("amount"))["total"] or Decimal("0")
        total_expenses = qs.filter(transaction_type="expense").aggregate(total=Sum("amount"))["total"] or Decimal("0")
        net_savings = total_income - total_expenses
        savings_rate = float(net_savings / total_income * 100) if total_income > 0 else 0.0

        # DataFrame for transactions
        frame = transaction_frame(request.user, month, year)
        if frame.empty:
            frame = pd.DataFrame(columns=["date", "category", "transaction_type", "amount", "description"])
        else:
            # Sort chronologically
            frame = frame.sort_values(by="date")

        ext = "xlsx" if export_format == "excel" else export_format
        filename = f"finance-report-{year}-{month:02d}.{ext}"

        # ----------------------------------------------------
        # 1. CSV EXPORT
        # ----------------------------------------------------
        if export_format == "csv":
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            # Add summary header lines
            writer = csv.writer(response)
            writer.writerow(["LEDGERLY PERSONAL FINANCE REPORT"])
            writer.writerow([f"Period: {month_name} {year}"])
            writer.writerow([f"Generated For: {request.user.username}"])
            writer.writerow([])
            writer.writerow(["SUMMARY METRICS"])
            writer.writerow(["Total Income", "Total Expenses", "Net Savings", "Savings Rate"])
            writer.writerow([float(total_income), float(total_expenses), float(net_savings), f"{savings_rate:.1f}%"])
            writer.writerow([])
            writer.writerow(["TRANSACTIONS LIST"])
            
            # Write pandas data
            frame.to_csv(response, index=False, mode='a')
            return response

        # ----------------------------------------------------
        # 2. EXCEL EXPORT (openpyxl)
        # ----------------------------------------------------
        if export_format == "excel":
            output = BytesIO()
            wb = openpyxl.Workbook()
            
            # Sheet 1: Summary Dashboard
            ws_summary = wb.active
            ws_summary.title = "Summary Dashboard"
            ws_summary.views.sheetView[0].showGridLines = True
            
            # Styling definitions
            font_title = Font(name="Segoe UI", size=16, bold=True, color="0f766e")
            font_section = Font(name="Segoe UI", size=12, bold=True, color="1e293b")
            font_header = Font(name="Segoe UI", size=11, bold=True, color="ffffff")
            font_bold = Font(name="Segoe UI", size=10, bold=True, color="1e293b")
            font_regular = Font(name="Segoe UI", size=10, color="334155")
            
            fill_teal = PatternFill(start_color="0f766e", end_color="0f766e", fill_type="solid")
            fill_indigo = PatternFill(start_color="4f46e5", end_color="4f46e5", fill_type="solid")
            fill_gray = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
            
            thin_border_side = Side(border_style="thin", color="cbd5e1")
            border_thin = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            double_bottom = Border(bottom=Side(style='double', color='0f766e'))
            
            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            align_right = Alignment(horizontal="right", vertical="center")

            # Title Block
            ws_summary["A1"] = "Ledgerly — Financial Statement Summary"
            ws_summary["A1"].font = font_title
            ws_summary["A2"] = f"Statement Period: {month_name} {year}  |  Account: {request.user.username}"
            ws_summary["A2"].font = Font(name="Segoe UI", size=10, italic=True, color="64748b")
            
            # KPI Cards
            ws_summary["A4"] = "Total Income"
            ws_summary["B4"] = "Total Expenses"
            ws_summary["C4"] = "Net Savings"
            ws_summary["D4"] = "Savings Rate"
            for col in ["A4", "B4", "C4", "D4"]:
                ws_summary[col].font = font_header
                ws_summary[col].fill = fill_indigo
                ws_summary[col].alignment = align_center
                ws_summary[col].border = border_thin
                
            ws_summary["A5"] = float(total_income)
            ws_summary["B5"] = float(total_expenses)
            ws_summary["C5"] = float(net_savings)
            ws_summary["D5"] = savings_rate / 100.0  # openpyxl percentage formatting will handle display
            
            ws_summary["A5"].number_format = '"Rs. "#,##0.00'
            ws_summary["B5"].number_format = '"Rs. "#,##0.00'
            ws_summary["C5"].number_format = '"Rs. "#,##0.00'
            ws_summary["D5"].number_format = '0.0%'
            
            for col in ["A5", "B5", "C5", "D5"]:
                ws_summary[col].font = font_bold
                ws_summary[col].alignment = align_center
                ws_summary[col].border = border_thin
                ws_summary[col].fill = fill_gray

            # Category Breakdown Title
            ws_summary["A7"] = "Expense Breakdown by Category"
            ws_summary["A7"].font = font_section
            
            # Category Breakdown Table
            ws_summary["A8"] = "Category"
            ws_summary["B8"] = "Amount Spent"
            ws_summary["C8"] = "% of Expenses"
            for col in ["A8", "B8", "C8"]:
                ws_summary[col].font = font_header
                ws_summary[col].fill = fill_teal
                ws_summary[col].alignment = align_left if col == "A8" else align_right
                ws_summary[col].border = border_thin
                
            categories_qs = qs.filter(transaction_type="expense").values("category").annotate(total=Sum("amount")).order_by("-total")
            row_idx = 9
            for cat in categories_qs:
                ws_summary[f"A{row_idx}"] = cat["category"]
                ws_summary[f"B{row_idx}"] = float(cat["total"])
                ws_summary[f"C{row_idx}"] = float(cat["total"] / total_expenses) if total_expenses > 0 else 0.0
                
                ws_summary[f"A{row_idx}"].font = font_regular
                ws_summary[f"B{row_idx}"].font = font_regular
                ws_summary[f"C{row_idx}"].font = font_regular
                
                ws_summary[f"B{row_idx}"].number_format = '"Rs. "#,##0.00'
                ws_summary[f"C{row_idx}"].number_format = '0.0%'
                
                ws_summary[f"A{row_idx}"].border = border_thin
                ws_summary[f"B{row_idx}"].border = border_thin
                ws_summary[f"C{row_idx}"].border = border_thin
                row_idx += 1
                
            # Budget Utilization Title
            ws_summary[f"A{row_idx + 1}"] = "Budget Threshold Review"
            ws_summary[f"A{row_idx + 1}"].font = font_section
            
            # Budget Table Headers
            start_budget_row = row_idx + 2
            ws_summary[f"A{start_budget_row}"] = "Category"
            ws_summary[f"B{start_budget_row}"] = "Limit Set"
            ws_summary[f"C{start_budget_row}"] = "Actual Spent"
            ws_summary[f"D{start_budget_row}"] = "Remaining"
            ws_summary[f"E{start_budget_row}"] = "Status"
            
            for col in ["A", "B", "C", "D", "E"]:
                cell = ws_summary[f"{col}{start_budget_row}"]
                cell.font = font_header
                cell.fill = fill_teal
                cell.alignment = align_left if col in ["A", "E"] else align_right
                cell.border = border_thin
                
            budgets_qs = Budget.objects.filter(user=request.user, month=month, year=year)
            bud_row_idx = start_budget_row + 1
            for b in budgets_qs:
                spent = qs.filter(transaction_type="expense", category__iexact=b.category).aggregate(total=Sum("amount"))["total"] or Decimal("0")
                remaining = b.monthly_limit - spent
                status_text = "OVER BUDGET" if spent > b.monthly_limit else "SAFE"
                
                ws_summary[f"A{bud_row_idx}"] = b.category
                ws_summary[f"B{bud_row_idx}"] = float(b.monthly_limit)
                ws_summary[f"C{bud_row_idx}"] = float(spent)
                ws_summary[f"D{bud_row_idx}"] = float(remaining)
                ws_summary[f"E{bud_row_idx}"] = status_text
                
                ws_summary[f"A{bud_row_idx}"].font = font_regular
                ws_summary[f"B{bud_row_idx}"].font = font_regular
                ws_summary[f"C{bud_row_idx}"].font = font_regular
                ws_summary[f"D{bud_row_idx}"].font = font_regular
                
                ws_summary[f"E{bud_row_idx}"].font = Font(name="Segoe UI", size=10, bold=True, color="991b1b" if spent > b.monthly_limit else "166534")
                
                ws_summary[f"B{bud_row_idx}"].number_format = '"Rs. "#,##0.00'
                ws_summary[f"C{bud_row_idx}"].number_format = '"Rs. "#,##0.00'
                ws_summary[f"D{bud_row_idx}"].number_format = '"Rs. "#,##0.00'
                
                for col in ["A", "B", "C", "D", "E"]:
                    ws_summary[f"{col}{bud_row_idx}"].border = border_thin
                bud_row_idx += 1

            # Auto-fit columns on summary page
            for col in ws_summary.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 15)

            # Sheet 2: Detailed Transactions
            ws_tx = wb.create_sheet("Transactions Detailed")
            ws_tx.views.sheetView[0].showGridLines = True
            
            # Write table columns
            headers = ["Date", "Category", "Type", "Amount", "Description"]
            for col_num, h in enumerate(headers, 1):
                cell = ws_tx.cell(row=1, column=col_num)
                cell.value = h
                cell.font = font_header
                cell.fill = fill_indigo
                cell.alignment = align_left if h != "Amount" else align_right
                cell.border = border_thin
                
            # Write data rows
            for row_idx, r in enumerate(frame.itertuples(index=False), 2):
                ws_tx.cell(row=row_idx, column=1, value=str(r.date)).alignment = align_left
                ws_tx.cell(row=row_idx, column=2, value=r.category).alignment = align_left
                
                type_cell = ws_tx.cell(row=row_idx, column=3, value=r.transaction_type.capitalize())
                type_cell.alignment = align_center
                type_cell.font = Font(name="Segoe UI", size=10, bold=True, color="166534" if r.transaction_type == "income" else "b91c1c")
                
                amt_cell = ws_tx.cell(row=row_idx, column=4, value=float(r.amount))
                amt_cell.alignment = align_right
                amt_cell.number_format = '"Rs. "#,##0.00'
                
                ws_tx.cell(row=row_idx, column=5, value=r.description or "-").alignment = align_left
                
                # Apply borders and styles
                for col_num in range(1, 6):
                    c = ws_tx.cell(row=row_idx, column=col_num)
                    c.border = border_thin
                    if col_num != 3: # Keep custom type color
                        c.font = font_regular
                    if row_idx % 2 == 0:
                        c.fill = fill_gray

            # Auto-fit columns on transactions page
            for col in ws_tx.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws_tx.column_dimensions[col_letter].width = max(max_len + 4, 15)

            wb.save(output)
            response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        # ----------------------------------------------------
        # 3. PDF EXPORT (ReportLab Platypus)
        # ----------------------------------------------------
        output = BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=letter,
            leftMargin=54,
            rightMargin=54,
            topMargin=54,
            bottomMargin=54
        )
        
        # Styles
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            name="ReportTitle",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=24,
            leading=28,
            textColor=colors.HexColor("#0f766e")
        )
        
        subtitle_style = ParagraphStyle(
            name="ReportSubTitle",
            parent=styles["Normal"],
            fontName="Helvetica-Oblique",
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#64748b")
        )
        
        section_style = ParagraphStyle(
            name="SectionHeading",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=13,
            leading=16,
            textColor=colors.HexColor("#1e293b"),
            spaceBefore=14,
            spaceAfter=6,
            keepWithNext=True
        )
        
        body_style = ParagraphStyle(
            name="ReportBody",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9,
            leading=13,
            textColor=colors.HexColor("#334155")
        )
        
        bold_body_style = ParagraphStyle(
            name="ReportBodyBold",
            parent=body_style,
            fontName="Helvetica-Bold"
        )
        
        insight_style = ParagraphStyle(
            name="InsightText",
            parent=body_style,
            fontSize=9,
            leading=13,
            textColor=colors.HexColor("#1e293b")
        )

        kpi_title_style = ParagraphStyle(
            name="KPITitle",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#475569"),
            alignment=1 # Centered
        )

        kpi_val_style = ParagraphStyle(
            name="KPIVal",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=13,
            leading=15,
            textColor=colors.HexColor("#0f172a"),
            alignment=1 # Centered
        )

        story = []

        # Branded Title Bar Table
        title_left = [
            Paragraph("LEDGERLY", title_style),
            Paragraph("Wealth Management Statement", subtitle_style)
        ]
        title_right = [
            Paragraph("<b>FINANCIAL REPORT</b>", ParagraphStyle(name="TR1", parent=body_style, fontName="Helvetica-Bold", fontSize=11, leading=14, alignment=2, textColor=colors.HexColor("#0f766e"))),
            Paragraph(f"Period: <b>{month_name} {year}</b>", ParagraphStyle(name="TR2", parent=body_style, alignment=2)),
            Paragraph(f"Account: <b>{request.user.username}</b>", ParagraphStyle(name="TR3", parent=body_style, alignment=2))
        ]
        
        title_table = Table([[title_left, title_right]], colWidths=[250, 254])
        title_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ]))
        story.append(title_table)
        
        # Horizontal accent bar
        accent_bar = Table([[""]], colWidths=[504], rowHeights=[3])
        accent_bar.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0f766e")),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(accent_bar)
        story.append(Spacer(1, 14))
        
        # 1. Executive Summary Block
        overview_text = (
            f"During the period of {month_name} {year}, your account recorded total inflows of "
            f"<b>{currency(total_income)}</b> and outbound expenses of <b>{currency(total_expenses)}</b>, "
            f"resulting in a net savings flow of <b>{currency(net_savings)}</b>. Below is a detailed breakdown "
            f"of your transactions, spending insights, and budget thresholds."
        )
        
        overview_table = Table(
            [[Paragraph("<b>Executive Overview</b>", ParagraphStyle(name="OverviewTitle", parent=section_style, spaceBefore=0, textColor=colors.HexColor("#0f766e")))],
             [Paragraph(overview_text, body_style)]],
            colWidths=[504]
        )
        overview_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
            ("PADDING", (0, 0), (-1, -1), 10),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
            ("LINELEFT", (0, 0), (-1, -1), 3, colors.HexColor("#0f766e")),
            ("TOPPADDING", (0, 0), (0, 0), 8),
            ("BOTTOMPADDING", (0, 0), (0, 0), 2),
        ]))
        story.append(overview_table)
        story.append(Spacer(1, 14))

        # 2. KPI Summary Cards Table (1 row, 4 columns)
        kpi_income_cell = [
            Paragraph("TOTAL INCOME", kpi_title_style),
            Spacer(1, 4),
            Paragraph(currency(total_income), ParagraphStyle(name="KpiIn", parent=kpi_val_style, textColor=colors.HexColor("#166534")))
        ]
        kpi_expense_cell = [
            Paragraph("TOTAL EXPENSES", kpi_title_style),
            Spacer(1, 4),
            Paragraph(currency(total_expenses), ParagraphStyle(name="KpiEx", parent=kpi_val_style, textColor=colors.HexColor("#991b1b")))
        ]
        kpi_savings_cell = [
            Paragraph("NET SAVINGS", kpi_title_style),
            Spacer(1, 4),
            Paragraph(currency(net_savings), ParagraphStyle(name="KpiSav", parent=kpi_val_style, textColor=colors.HexColor("#075985")))
        ]
        kpi_rate_cell = [
            Paragraph("SAVINGS RATE", kpi_title_style),
            Spacer(1, 4),
            Paragraph(f"{savings_rate:.1f}%", ParagraphStyle(name="KpiRate", parent=kpi_val_style, textColor=colors.HexColor("#6b21a8")))
        ]
        
        kpi_table = Table([[kpi_income_cell, kpi_expense_cell, kpi_savings_cell, kpi_rate_cell]], colWidths=[126, 126, 126, 126])
        kpi_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("PADDING", (0, 0), (-1, -1), 8),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            
            # Colored background bands
            ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#f0fdf4")),
            ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#fff1f2")),
            ("BACKGROUND", (2, 0), (2, 0), colors.HexColor("#f0f9ff")),
            ("BACKGROUND", (3, 0), (3, 0), colors.HexColor("#fdf4ff")),
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 14))

        # 3. Personalized Financial Insights (Callout Feed)
        story.append(Paragraph("Personalized Financial Insights", section_style))
        insights = financial_insights(request.user, month, year)
        insights_data = []
        for ins in insights:
            bullet = Paragraph("💡", ParagraphStyle(name="Bullet", parent=body_style, fontSize=12, alignment=1))
            text = Paragraph(ins, insight_style)
            insights_data.append([bullet, text])
            
        insights_table = Table(insights_data, colWidths=[24, 480])
        insights_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor("#f1f5f9")),
        ]))
        story.append(insights_table)
        story.append(Spacer(1, 14))

        # 4. Budget Utilization Grid with Progress Bars
        budgets_list = Budget.objects.filter(user=request.user, month=month, year=year)
        if budgets_list.exists():
            story.append(Paragraph("Category Budget Compliance", section_style))
            budget_data = [["Category", "Budget Limit", "Amount Spent", "Remaining", "Utilization Bar", "Status"]]
            
            for b in budgets_list:
                spent = qs.filter(transaction_type="expense", category__iexact=b.category).aggregate(total=Sum("amount"))["total"] or Decimal("0")
                remaining = b.monthly_limit - spent
                pct = float(spent / b.monthly_limit * 100) if b.monthly_limit > 0 else 0.0
                bar_graphic = make_progress_bar(pct)
                status_label = Paragraph(
                    f"<font color='#b91c1c'><b>OVER ({pct:.0f}%)</b></font>" if spent > b.monthly_limit 
                    else f"<font color='#15803d'><b>SAFE ({pct:.0f}%)</b></font>",
                    body_style
                )
                
                budget_data.append([
                    Paragraph(b.category, body_style),
                    Paragraph(currency(b.monthly_limit), body_style),
                    Paragraph(currency(spent), body_style),
                    Paragraph(currency(remaining), body_style),
                    bar_graphic,
                    status_label
                ])
                
            budget_table = Table(budget_data, colWidths=[100, 80, 80, 80, 94, 70])
            budget_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("PADDING", (0, 0), (-1, -1), 6),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ]))
            story.append(budget_table)
            story.append(Spacer(1, 14))

        # 5. Detailed Transactions Log (with autowrap description Paragraph)
        story.append(Paragraph("Detailed Transactions Statement", section_style))
        tx_data = [["Date", "Category", "Description", "Type", "Amount"]]
        
        for r in frame.itertuples(index=False):
            type_label = Paragraph(
                "<font color='#166534'><b>Income</b></font>" if r.transaction_type == "income" 
                else "<font color='#991b1b'><b>Expense</b></font>",
                body_style
            )
            
            # Wrap description in a paragraph to force auto-wrap
            desc_text = r.description if (r.description and str(r.description).strip() != "") else "-"
            desc_para = Paragraph(desc_text, ParagraphStyle(name="DescWrap", parent=body_style, fontSize=8, leading=10))
            
            tx_data.append([
                Paragraph(str(r.date), body_style),
                Paragraph(r.category, body_style),
                desc_para,
                type_label,
                Paragraph(currency(r.amount), bold_body_style if r.transaction_type == "income" else body_style)
            ])
            
        tx_table = Table(tx_data, colWidths=[64, 90, 190, 60, 100], repeatRows=1)
        tx_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("PADDING", (0, 0), (-1, -1), 6),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ]))
        story.append(tx_table)

        # Build document
        doc.build(story, canvasmaker=NumberedCanvas)
        response = HttpResponse(output.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
